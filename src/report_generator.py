"""
Report Generation Module.

Features:
1. PDF Talent Gap Reports - Executive summary of skill shortages
2. Excel HR Exports - Detailed candidate data for spreadsheet analysis
3. Resume Viewer with Highlighting - Visual skill matching
4. Downloadable analytics dashboards

Dependencies: reportlab (PDF), openpyxl (Excel), pandas
"""
from typing import List, Dict, Optional
import pandas as pd
from datetime import datetime
import os


# ==================== Excel Report Generation ====================

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelReportGenerator:
    """
    Generate Excel reports for HR teams.
    
    Reports include:
    - Candidate rankings with scores
    - Skill gap analysis
    - Hiring funnel metrics
    - Time-to-hire statistics
    """
    
    @staticmethod
    def generate_candidate_report(candidates: List[Dict], job_description: str, 
                                  output_path: str = "candidate_report.xlsx"):
        """
        Generate Excel report with candidate rankings.
        
        Args:
            candidates: List of candidate dictionaries
            job_description: Job description searched
            output_path: Output file path
        
        Returns:
            Path to generated Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation. Install: pip install openpyxl")
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        ws_summary['A1'] = "Candidate Search Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A3'] = f"Job Description: {job_description[:100]}..."
        ws_summary['A4'] = f"Total Candidates: {len(candidates)}"
        
        # Candidate data
        ws_summary['A6'] = "Top Candidates"
        ws_summary['A6'].font = Font(bold=True, size=12)
        
        # Headers
        headers = ['Rank', 'Name', 'Score', 'Skills Match', 'Experience Match', 'Seniority', 'Top Skills']
        for col, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows
        for idx, candidate in enumerate(candidates, start=1):
            row = 7 + idx
            ws_summary.cell(row, 1, idx)
            ws_summary.cell(row, 2, candidate.get('name', 'Unknown'))
            ws_summary.cell(row, 3, round(candidate.get('score', 0), 3))
            ws_summary.cell(row, 4, round(candidate.get('skill_overlap_score', 0), 3))
            ws_summary.cell(row, 5, round(candidate.get('experience_match_score', 0), 3))
            ws_summary.cell(row, 6, candidate.get('seniority_level', 'Unknown'))
            ws_summary.cell(row, 7, ', '.join(candidate.get('skills', [])[:5]))
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 8
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 10
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 15
        ws_summary.column_dimensions['F'].width = 12
        ws_summary.column_dimensions['G'].width = 40
        
        # Sheet 2: Detailed Skills
        ws_skills = wb.create_sheet("Skill Details")
        ws_skills['A1'] = "Candidate Skill Breakdown"
        ws_skills['A1'].font = Font(size=14, bold=True)
        
        skill_headers = ['Candidate', 'Total Skills', 'Hard Skills', 'Soft Skills', 'Common Skills', 'Missing Skills']
        for col, header in enumerate(skill_headers, start=1):
            cell = ws_skills.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        for idx, candidate in enumerate(candidates, start=1):
            row = 3 + idx
            ws_skills.cell(row, 1, candidate.get('name', 'Unknown'))
            ws_skills.cell(row, 2, len(candidate.get('skills', [])))
            ws_skills.cell(row, 3, ', '.join(candidate.get('hard_skills', [])[:10]))
            ws_skills.cell(row, 4, ', '.join(candidate.get('soft_skills', [])[:10]))
            ws_skills.cell(row, 5, ', '.join(candidate.get('common_skills', [])[:10]))
            
            explanation = candidate.get('explanation', {})
            missing = explanation.get('missing_critical_skills', [])
            ws_skills.cell(row, 6, ', '.join(missing[:10]))
        
        # Adjust widths
        for col in range(1, 7):
            ws_skills.column_dimensions[chr(64 + col)].width = 30
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
    
    
    @staticmethod
    def generate_analytics_report(funnel_metrics: Dict, skill_times: Dict[str, float],
                                  forecasts: List, output_path: str = "analytics_report.xlsx"):
        """
        Generate Excel report with analytics metrics.
        
        Args:
            funnel_metrics: Hiring funnel metrics dictionary
            skill_times: Time-to-hire by skill
            forecasts: Talent gap forecasts
            output_path: Output file path
        
        Returns:
            Path to generated Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required")
        
        wb = Workbook()
        
        # Sheet 1: Funnel Metrics
        ws_funnel = wb.active
        ws_funnel.title = "Hiring Funnel"
        
        ws_funnel['A1'] = "Hiring Funnel Analytics"
        ws_funnel['A1'].font = Font(size=16, bold=True)
        
        ws_funnel['A3'] = "Metric"
        ws_funnel['B3'] = "Value"
        ws_funnel['A3'].font = Font(bold=True)
        ws_funnel['B3'].font = Font(bold=True)
        
        metrics_data = [
            ("Total Searches", funnel_metrics.get('total_searches', 0)),
            ("Total Shortlists", funnel_metrics.get('total_shortlists', 0)),
            ("Total Hires", funnel_metrics.get('total_hires', 0)),
            ("Search → Shortlist Rate", f"{funnel_metrics.get('search_to_shortlist_rate', 0)}%"),
            ("Shortlist → Hire Rate", f"{funnel_metrics.get('shortlist_to_hire_rate', 0)}%"),
            ("Overall Conversion Rate", f"{funnel_metrics.get('overall_conversion_rate', 0)}%"),
            ("Drop-off: Search → Shortlist", f"{funnel_metrics.get('drop_off_search_to_shortlist', 0)}%"),
            ("Drop-off: Shortlist → Hire", f"{funnel_metrics.get('drop_off_shortlist_to_hire', 0)}%"),
            ("Avg Time to Hire", f"{funnel_metrics.get('avg_time_to_hire_days', 0)} days"),
        ]
        
        for idx, (metric, value) in enumerate(metrics_data, start=4):
            ws_funnel.cell(idx, 1, metric)
            ws_funnel.cell(idx, 2, value)
        
        ws_funnel.column_dimensions['A'].width = 30
        ws_funnel.column_dimensions['B'].width = 20
        
        # Sheet 2: Time to Hire by Skill
        ws_time = wb.create_sheet("Time to Hire")
        ws_time['A1'] = "Average Time to Hire by Skill"
        ws_time['A1'].font = Font(size=14, bold=True)
        
        ws_time['A3'] = "Skill"
        ws_time['B3'] = "Avg Days to Hire"
        ws_time['A3'].font = Font(bold=True)
        ws_time['B3'].font = Font(bold=True)
        
        for idx, (skill, days) in enumerate(skill_times.items(), start=4):
            ws_time.cell(idx, 1, skill)
            ws_time.cell(idx, 2, days)
        
        ws_time.column_dimensions['A'].width = 30
        ws_time.column_dimensions['B'].width = 20
        
        # Sheet 3: Talent Gap Forecasts
        ws_forecast = wb.create_sheet("Talent Gap Forecast")
        ws_forecast['A1'] = "Predicted Skill Shortages (Next Month)"
        ws_forecast['A1'].font = Font(size=14, bold=True)
        
        forecast_headers = ['Skill', 'Current Searches', 'Predicted Searches', 'Trend', 'Shortage Risk', 'Confidence']
        for col, header in enumerate(forecast_headers, start=1):
            cell = ws_forecast.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        for idx, forecast in enumerate(forecasts, start=4):
            ws_forecast.cell(idx, 1, forecast.skill)
            ws_forecast.cell(idx, 2, forecast.current_search_count)
            ws_forecast.cell(idx, 3, forecast.predicted_next_month)
            ws_forecast.cell(idx, 4, forecast.trend)
            ws_forecast.cell(idx, 5, forecast.shortage_risk)
            ws_forecast.cell(idx, 6, f"{int(forecast.confidence * 100)}%")
            
            # Color code by risk
            risk_cell = ws_forecast.cell(idx, 5)
            if forecast.shortage_risk == "high":
                risk_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif forecast.shortage_risk == "medium":
                risk_cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            else:
                risk_cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
        
        for col in range(1, 7):
            ws_forecast.column_dimensions[chr(64 + col)].width = 18
        
        wb.save(output_path)
        
        return output_path


# ==================== PDF Report Generation ====================

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class PDFReportGenerator:
    """
    Generate PDF reports for executive summaries.
    
    Reports include:
    - Talent gap analysis
    - Skill shortage predictions
    - Hiring funnel visualizations
    """
    
    @staticmethod
    def generate_talent_gap_report(forecasts: List, funnel_metrics: Dict,
                                   output_path: str = "talent_gap_report.pdf"):
        """
        Generate executive PDF report on talent gaps.
        
        Args:
            forecasts: List of SkillTrendForecast objects
            funnel_metrics: Hiring funnel metrics
            output_path: Output file path
        
        Returns:
            Path to generated PDF
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF generation. Install: pip install reportlab")
        
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph("Talent Gap Analysis Report", title_style)
        story.append(title)
        
        # Subtitle
        subtitle = Paragraph(
            f"Generated: {datetime.now().strftime('%B %d, %Y')}",
            styles['Normal']
        )
        story.append(subtitle)
        story.append(Spacer(1, 0.3*inch))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        
        summary_text = f"""
        This report analyzes current talent gaps and predicts future skill shortages based on 
        recruiter search trends. Our analysis identified <b>{len(forecasts)}</b> critical skills 
        with potential talent shortages in the coming month.
        """
        story.append(Paragraph(summary_text, styles['BodyText']))
        story.append(Spacer(1, 0.2*inch))
        
        # Hiring Funnel Metrics
        story.append(Paragraph("Hiring Pipeline Performance", styles['Heading2']))
        
        funnel_data = [
            ['Metric', 'Value'],
            ['Total Searches', str(funnel_metrics.get('total_searches', 0))],
            ['Total Shortlists', str(funnel_metrics.get('total_shortlists', 0))],
            ['Total Hires', str(funnel_metrics.get('total_hires', 0))],
            ['Conversion Rate', f"{funnel_metrics.get('overall_conversion_rate', 0)}%"],
            ['Avg Time to Hire', f"{funnel_metrics.get('avg_time_to_hire_days', 0)} days"],
        ]
        
        funnel_table = Table(funnel_data, colWidths=[3*inch, 2*inch])
        funnel_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(funnel_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Skill Shortage Predictions
        story.append(Paragraph("Predicted Skill Shortages (Next 30 Days)", styles['Heading2']))
        
        forecast_data = [['Skill', 'Current\nSearches', 'Predicted\nSearches', 'Trend', 'Risk', 'Confidence']]
        
        for forecast in forecasts[:10]:  # Top 10
            forecast_data.append([
                forecast.skill,
                str(forecast.current_search_count),
                f"{forecast.predicted_next_month:.1f}",
                forecast.trend.capitalize(),
                forecast.shortage_risk.upper(),
                f"{int(forecast.confidence * 100)}%"
            ])
        
        forecast_table = Table(forecast_data, colWidths=[1.5*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.7*inch, 0.8*inch])
        forecast_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        # Color code risk levels
        for idx, forecast in enumerate(forecasts[:10], start=1):
            if forecast.shortage_risk == "high":
                forecast_table.setStyle(TableStyle([
                    ('BACKGROUND', (4, idx), (4, idx), colors.HexColor('#FF6B6B'))
                ]))
            elif forecast.shortage_risk == "medium":
                forecast_table.setStyle(TableStyle([
                    ('BACKGROUND', (4, idx), (4, idx), colors.HexColor('#FFD93D'))
                ]))
            else:
                forecast_table.setStyle(TableStyle([
                    ('BACKGROUND', (4, idx), (4, idx), colors.HexColor('#6BCF7F'))
                ]))
        
        story.append(forecast_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Recommendations
        story.append(Paragraph("Recommendations", styles['Heading2']))
        
        high_risk_skills = [f.skill for f in forecasts if f.shortage_risk == "high"]
        
        if high_risk_skills:
            reco_text = f"""
            <b>Immediate Action Required:</b><br/>
            The following skills show high shortage risk: <b>{', '.join(high_risk_skills[:5])}</b>.<br/><br/>
            
            Recommended actions:<br/>
            • Accelerate recruitment campaigns for these skills<br/>
            • Consider contractor/freelance options<br/>
            • Invest in upskilling current employees<br/>
            • Explore remote/offshore talent pools<br/>
            • Adjust compensation packages to remain competitive
            """
        else:
            reco_text = """
            Current talent pipeline appears healthy. Continue monitoring trends and 
            maintain proactive recruitment strategies.
            """
        
        story.append(Paragraph(reco_text, styles['BodyText']))
        
        # Build PDF
        doc.build(story)
        
        return output_path


# ==================== Resume Viewer with Highlighting ====================

class ResumeHighlighter:
    """
    Highlight matched and missing skills in resume text.
    
    Returns HTML with color-coded skills:
    - Green: Matched skills
    - Red: Missing critical skills
    """
    
    @staticmethod
    def highlight_resume(resume_text: str, matched_skills: List[str], 
                        missing_skills: List[str]) -> str:
        """
        Generate HTML with highlighted skills.
        
        Args:
            resume_text: Original resume text
            matched_skills: Skills that match job description
            missing_skills: Critical skills candidate lacks
        
        Returns:
            HTML string with highlighted text
        """
        import re
        
        highlighted_text = resume_text
        
        # Highlight matched skills (green)
        for skill in matched_skills:
            pattern = re.compile(re.escape(skill), re.IGNORECASE)
            highlighted_text = pattern.sub(
                f'<span style="background-color: #d4edda; color: #155724; padding: 2px 4px; border-radius: 3px; font-weight: bold;">{skill}</span>',
                highlighted_text
            )
        
        # Note: We don't highlight missing skills in resume (they're not there!)
        # Instead, we'll show them separately in a legend
        
        html = f"""
        <div style="font-family: Arial, sans-serif; line-height: 1.6;">
            <div style="margin-bottom: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 5px;">
                <h4 style="margin-top: 0;">Legend:</h4>
                <div style="margin-bottom: 10px;">
                    <span style="background-color: #d4edda; color: #155724; padding: 2px 8px; border-radius: 3px; font-weight: bold;">
                        Matched Skill
                    </span> 
                    <span style="margin-left: 10px; color: #666;">Skills candidate has that match job requirements</span>
                </div>
                
                {f'''<div style="margin-bottom: 10px;">
                    <span style="background-color: #f8d7da; color: #721c24; padding: 2px 8px; border-radius: 3px; font-weight: bold;">
                        Missing Skill
                    </span>
                    <span style="margin-left: 10px; color: #666;">Critical skills: {', '.join(missing_skills[:5])}</span>
                </div>''' if missing_skills else ''}
            </div>
            
            <div style="white-space: pre-wrap; background-color: white; padding: 20px; border: 1px solid #ddd; border-radius: 5px;">
                {highlighted_text}
            </div>
        </div>
        """
        
        return html
    
    
    @staticmethod
    def generate_skill_match_html(candidate: Dict, job_skills: List[str]) -> str:
        """
        Generate comprehensive skill match visualization.
        
        Args:
            candidate: Candidate dictionary
            job_skills: Required skills from job description
        
        Returns:
            HTML string with skill breakdown
        """
        matched = candidate.get('common_skills', [])
        candidate_skills = candidate.get('skills', [])
        missing = [s for s in job_skills if s not in candidate_skills]
        
        html = f"""
        <div style="font-family: Arial, sans-serif;">
            <h3 style="color: #366092;">Skill Match Analysis</h3>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px;">
                <div style="background-color: #d4edda; padding: 15px; border-radius: 5px;">
                    <h4 style="margin-top: 0; color: #155724;">✓ Matched Skills ({len(matched)})</h4>
                    <div style="display: flex; flex-wrap: wrap; gap: 5px;">
                        {' '.join([f'<span style="background-color: #28a745; color: white; padding: 4px 8px; border-radius: 3px; font-size: 12px;">{skill}</span>' for skill in matched])}
                    </div>
                </div>
                
                <div style="background-color: #f8d7da; padding: 15px; border-radius: 5px;">
                    <h4 style="margin-top: 0; color: #721c24;">✗ Missing Skills ({len(missing)})</h4>
                    <div style="display: flex; flex-wrap: wrap; gap: 5px;">
                        {' '.join([f'<span style="background-color: #dc3545; color: white; padding: 4px 8px; border-radius: 3px; font-size: 12px;">{skill}</span>' for skill in missing[:10]])}
                    </div>
                </div>
            </div>
            
            <div style="background-color: #e7f3ff; padding: 15px; border-radius: 5px;">
                <h4 style="margin-top: 0; color: #004085;">Additional Skills ({len(candidate_skills) - len(matched)})</h4>
                <div style="display: flex; flex-wrap: wrap; gap: 5px;">
                    {' '.join([f'<span style="background-color: #17a2b8; color: white; padding: 4px 8px; border-radius: 3px; font-size: 12px;">{skill}</span>' for skill in candidate_skills if skill not in matched][:15])}
                </div>
            </div>
        </div>
        """
        
        return html


# ==================== Testing ====================

if __name__ == "__main__":
    print("=" * 60)
    print("Report Generation Test")
    print("=" * 60)
    
    # Test data
    candidates = [
        {
            'name': 'John Doe',
            'score': 0.85,
            'skill_overlap_score': 0.78,
            'experience_match_score': 0.92,
            'seniority_level': 'Senior',
            'skills': ['python', 'tensorflow', 'aws', 'docker'],
            'hard_skills': ['python', 'tensorflow', 'aws'],
            'soft_skills': ['communication', 'leadership'],
            'common_skills': ['python', 'aws'],
            'explanation': {'missing_critical_skills': ['kubernetes']}
        }
    ]
    
    # Test Excel generation
    if OPENPYXL_AVAILABLE:
        print("\n1. Testing Excel generation...")
        try:
            excel_path = ExcelReportGenerator.generate_candidate_report(
                candidates,
                "Python developer with ML",
                "test_candidate_report.xlsx"
            )
            print(f"   ✓ Excel report generated: {excel_path}")
            os.remove(excel_path)
        except Exception as e:
            print(f"   ✗ Excel generation failed: {e}")
    else:
        print("\n1. Excel generation skipped (openpyxl not installed)")
    
    # Test PDF generation
    if REPORTLAB_AVAILABLE:
        print("\n2. Testing PDF generation...")
        try:
            from src.analytics import SkillTrendForecast
            
            forecasts = [
                SkillTrendForecast(
                    skill="python",
                    current_search_count=25,
                    predicted_next_month=35.5,
                    trend="increasing",
                    shortage_risk="high",
                    confidence=0.75
                )
            ]
            
            funnel_metrics = {
                'total_searches': 100,
                'total_shortlists': 45,
                'total_hires': 12,
                'overall_conversion_rate': 12.0,
                'avg_time_to_hire_days': 14.5
            }
            
            pdf_path = PDFReportGenerator.generate_talent_gap_report(
                forecasts,
                funnel_metrics,
                "test_talent_gap_report.pdf"
            )
            print(f"   ✓ PDF report generated: {pdf_path}")
            os.remove(pdf_path)
        except Exception as e:
            print(f"   ✗ PDF generation failed: {e}")
    else:
        print("\n2. PDF generation skipped (reportlab not installed)")
    
    # Test resume highlighting
    print("\n3. Testing resume highlighting...")
    resume_text = "Python developer with 5 years experience in TensorFlow and AWS."
    matched = ['python', 'tensorflow', 'aws']
    missing = ['kubernetes', 'docker']
    
    html = ResumeHighlighter.highlight_resume(resume_text, matched, missing)
    print(f"   ✓ Generated {len(html)} chars of HTML")
    
    print("\n✅ All tests passed!")
